import os
# admin_handlers.py

import re
from datetime import datetime, timedelta

from telegram import Update
from telegram.ext import ContextTypes

from config import FILIALS, ADMIN_IDS
from database import (
    get_all_users, get_incomes_by_filial_and_period, get_expenses_by_filial_and_period,
    admin_approve_income, admin_reject_income,
    admin_approve_expense, admin_reject_expense,
    get_income_by_id, get_expense_by_id,
    get_user_by_id, get_connection, delete_expense, delete_income
)
from keyboards import (
    admin_panel_keyboard, filial_inline_keyboard, period_select_keyboard,
    back_keyboard, approval_inline_keyboard, export_inline_keyboard, malumotlar_period_keyboard
)
from states import State

# admin_handlers.py

# admin_handlers.py

async def admin_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "Orqaga":
        await update.message.reply_text("Asosiy menyuga qaytish",
                                        reply_markup=context.user_data["main_menu_keyboard"])
        return State.MAIN_MENU

    if text == "Daromad va xarajatlar":
        await update.message.reply_text("Qaysi filialni ko'rmoqchisiz?",
                                        reply_markup=filial_inline_keyboard())
        return State.ADMIN_SELECT_FILIAL

    if text == "Userlar":
        # User management
        users = get_all_users()
        msg = "Barcha userlar:\n"
        for u in users:
            # (id, tg_id, filial_id, username, full_name, is_admin)
            role = "Admin" if u[5] == 1 else "User"
            msg += f"ID: {u[0]}, TG: {u[1]}, Filial: {FILIALS.get(u[2], 'Nomalum')}, Role: {role}, Name: {u[4]}\n"
        await update.message.reply_text(msg, reply_markup=admin_panel_keyboard())
        return State.USER_MANAGEMENT

    if text == "Malumotlar":  # Updated from "Bugungi ma'lumotlar"
        await update.message.reply_text("Filialni tanlang:", reply_markup=filial_inline_keyboard())
        return State.ADMIN_SELECT_TODAY_FILIAL  # Transition to new state

    if text == "Export":
        await update.message.reply_text("Filialni tanlang:", reply_markup=filial_inline_keyboard())
        return State.ADMIN_EXPORT_FILIAL

    await update.message.reply_text("Noma'lum buyruq", reply_markup=admin_panel_keyboard())
    return State.ADMIN_MENU



# admin_handlers.py
async def admin_select_filial_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    match = re.match(r"filial_(\d+)", data)
    if match:
        filial_id = int(match.group(1))
        context.user_data["selected_filial"] = filial_id
        context.user_data.pop("selected_today_filial", None)  # Malumotlar bo'limi tanlovini tozalash
        await query.message.reply_text("Davrni tanlang:", reply_markup=period_select_keyboard())
        await query.answer()
        return State.ADMIN_SELECT_PERIOD
    else:
        await query.message.reply_text("Filial topilmadi.", reply_markup=admin_panel_keyboard())
        await query.answer()
        return State.ADMIN_MENU

# admin_select_today_filial_callback
async def admin_select_today_filial_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    match = re.match(r"filial_(\d+)", data)
    if match:
        filial_id = int(match.group(1))
        context.user_data["selected_today_filial"] = filial_id
        context.user_data.pop("selected_filial", None)  # Daromad va xarajatlar bo'limi tanlovini tozalash
        await query.message.edit_reply_markup(reply_markup=None)  # Inline tugmalarni olib tashlash
        await query.message.reply_text("Davrni tanlang:", reply_markup=malumotlar_period_keyboard())
        await query.answer()
        return State.ADMIN_SELECT_TODAY_PERIOD  # Yangi holat
    else:
        await query.message.reply_text("Filial topilmadi.", reply_markup=admin_panel_keyboard())
        await query.answer()
        return State.ADMIN_MENU

# admin_handlers.py

# admin_handlers.py

from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ContextTypes


async def admin_select_period_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    selected_filial = context.user_data.get("selected_filial")
    selected_today_filial = context.user_data.get("selected_today_filial")

    if not selected_filial and not selected_today_filial:
        await update.message.reply_text("Filial tanlanmagan!",
                                        reply_markup=admin_panel_keyboard())
        return State.ADMIN_MENU

    if text == "Orqaga":
        await update.message.reply_text("Admin menyu", reply_markup=admin_panel_keyboard())
        return State.ADMIN_MENU

    today = datetime.now()

    if text == "Bugungi":
        start_date = today.strftime("%Y-%m-%d")
        end_date = start_date
    elif text == "Kechagi":
        d = today - timedelta(days=1)
        start_date = d.strftime("%Y-%m-%d")
        end_date = start_date
    elif text == "1-Haftalik":
        d = today - timedelta(days=7)
        start_date = d.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif text == "1-Oylik":
        d = today - timedelta(days=30)
        start_date = d.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif text == "Davrni tanlang":
        await update.message.reply_text(
            "Iltimos, davrning boshlanish sanasini quyidagi formatda kiriting: YYYY-MM-DD",
            reply_markup=ReplyKeyboardMarkup([["Orqaga"]], resize_keyboard=True)
        )
        context.user_data["awaiting_start_date"] = True
        return State.ADMIN_SELECT_CUSTOM_PERIOD
    else:
        await update.message.reply_text("Noma'lum tanlov.",
                                        reply_markup=period_select_keyboard())
        return State.ADMIN_SELECT_PERIOD

    if selected_filial:
        await show_filial_report(update, context, selected_filial, start_date, end_date)
    elif selected_today_filial:
        await show_pending_approvals(update, context, selected_today_filial, start_date, end_date)

    return State.ADMIN_MENU

async def admin_select_custom_period_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text

    if context.user_data.get("awaiting_start_date"):
        try:
            start_date = datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
            context.user_data["custom_start_date"] = start_date
            context.user_data["awaiting_start_date"] = False

            await update.message.reply_text(
                "Endi davrning tugash sanasini kiriting: YYYY-MM-DD",
                reply_markup=ReplyKeyboardMarkup([["Orqaga"]], resize_keyboard=True)
            )
            context.user_data["awaiting_end_date"] = True
            return State.ADMIN_SELECT_CUSTOM_PERIOD

        except ValueError:
            await update.message.reply_text(
                "Sanani noto'g'ri formatda kiritdingiz. Iltimos, YYYY-MM-DD formatidan foydalaning.",
                reply_markup=ReplyKeyboardMarkup([["Orqaga"]], resize_keyboard=True)
            )
            return State.ADMIN_SELECT_CUSTOM_PERIOD

    if context.user_data.get("awaiting_end_date"):
        try:
            end_date = datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
            context.user_data["custom_end_date"] = end_date
            context.user_data["awaiting_end_date"] = False

            start_date = context.user_data.get("custom_start_date")
            selected_filial = context.user_data.get("selected_filial")
            selected_today_filial = context.user_data.get("selected_today_filial")

            if selected_filial:
                await show_filial_report(update, context, selected_filial, start_date, end_date)
            elif selected_today_filial:
                await show_pending_approvals(update, context, selected_today_filial, start_date, end_date)

            return State.ADMIN_MENU

        except ValueError:
            await update.message.reply_text(
                "Sanani noto'g'ri formatda kiritdingiz. Iltimos, YYYY-MM-DD formatidan foydalaning.",
                reply_markup=ReplyKeyboardMarkup([["Orqaga"]], resize_keyboard=True)
            )
            return State.ADMIN_SELECT_CUSTOM_PERIOD

    await update.message.reply_text("Noma'lum xatolik yuz berdi.",
                                    reply_markup=period_select_keyboard())
    return State.ADMIN_SELECT_PERIOD


# admin_handlers.py
async def admin_enter_period_manual_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "Orqaga":
        await update.message.reply_text("Davrni tanlang:", reply_markup=period_select_keyboard())
        return State.ADMIN_SELECT_PERIOD

    match = re.match(r"(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})", text)
    if not match:
        await update.message.reply_text("Noto'g'ri format. Masalan: 2025-01-01 - 2025-01-07",
                                        reply_markup=back_keyboard())
        return State.ADMIN_ENTER_PERIOD_MANUAL

    start_date, end_date = match.group(1), match.group(2)

    # Filial tanlovini aniqlash
    if "selected_filial" in context.user_data:
        filial_id = context.user_data.get("selected_filial")
        await show_filial_report(update, context, filial_id, start_date, end_date)
        return State.ADMIN_VIEW_REPORT
    elif "selected_today_filial" in context.user_data:
        filial_id = context.user_data.get("selected_today_filial")
        await show_pending_approvals(update, context, filial_id, start_date, end_date)
        return State.ADMIN_MENU
    else:
        await update.message.reply_text("Filial tanlanmagan!",
                                        reply_markup=admin_panel_keyboard())
        return State.ADMIN_MENU




# admin_handlers.py

async def show_pending_approvals(update: Update, context: ContextTypes.DEFAULT_TYPE, filial_id, start_date, end_date):
    conn = get_connection()
    cursor = conn.cursor()

    # Daromadlar uchun tasdiqlash kutilmoqda
    cursor.execute("""
        SELECT * FROM incomes
        WHERE filial_id=? AND date_kiritilgan BETWEEN ? AND ? AND confirm_status=1 AND admin_approval=0
    """, (filial_id, start_date, end_date))
    pending_incomes = cursor.fetchall()

    # Xarajatlar uchun tasdiqlash kutilmoqda
    cursor.execute("""
        SELECT * FROM expenses
        WHERE filial_id=? AND date_kiritilgan BETWEEN ? AND ? AND confirm_status=1 AND admin_approval=0
    """, (filial_id, start_date, end_date))
    pending_expenses = cursor.fetchall()
    conn.close()

    if not pending_incomes and not pending_expenses:
        await update.message.reply_text("Hozircha tasdiqlash yoki Rad etish uchun kutilayotgan ma'lumot yo‘q.",
                                        reply_markup=admin_panel_keyboard())
        return

    # Tasdiqlanmagan daromadlar
    if pending_incomes:
        await update.message.reply_text("Tasdiqlanmagan daromad(lar):")
        for inc in pending_incomes:
            inc_id = inc[0]
            user_id = inc[1]
            user_row = get_user_by_id(user_id)
            user_name = user_row[4] if user_row else "Noma'lum user"

            s_naqt, s_humo, s_uzkard, s_clic, s_payme = inc[4], inc[5], inc[6], inc[7], inc[8]
            total_sum = s_naqt + s_humo + s_uzkard + s_clic + s_payme

            txt = (f"#DaromadID {inc_id}  (User: {user_name})\n"
                   f"Naqt: {s_naqt}, Humo: {s_humo}, Uzcard: {s_uzkard}, Clic: {s_clic}, Payme: {s_payme}\n"
                   f"Umumiy: {total_sum}")
            await update.message.reply_text(
                txt,
                reply_markup=approval_inline_keyboard("income", inc_id)
            )
    else:
        await update.message.reply_text("Daromad tasdiqlanmagan yo‘q.")

    # Tasdiqlanmagan xarajatlar
    if pending_expenses:
        await update.message.reply_text("Tasdiqlanmagan xarajat(lar):")
        for xp in pending_expenses:
            xp_id = xp[0]
            user_id = xp[1]
            user_row = get_user_by_id(user_id)
            user_name = user_row[4] if user_row else "Noma'lum user"

            s_summa = xp[4]
            s_info = xp[5]
            txt = (f"#XarajatID {xp_id} (User: {user_name})\n"
                   f"Summa: {s_summa} so'm, Info: {s_info}")
            await update.message.reply_text(
                txt,
                reply_markup=approval_inline_keyboard("expense", xp_id)
            )
    else:
        await update.message.reply_text("Xarajat tasdiqlanmagan yo‘q.")

    await update.message.reply_text("Quyidagilardan tasdiqlang yoki rad eting.",
                                    reply_markup=admin_panel_keyboard())

# admin_handlers.py

# admin_handlers.py

# admin_handlers.py

# admin_handlers.py

async def show_filial_report(update: Update, context: ContextTypes.DEFAULT_TYPE,
                             filial_id, start_date, end_date):
    incomes = get_incomes_by_filial_and_period(filial_id, start_date, end_date)
    expenses = get_expenses_by_filial_and_period(filial_id, start_date, end_date)

    if not incomes and not expenses:
        msg = "Ma'lumot topilmadi."
        keyboard = admin_panel_keyboard()
        await update.effective_message.reply_text(
            msg,
            parse_mode='Markdown',
            reply_markup=keyboard
        )
        return

    # Organize incomes and expenses by date
    income_by_date = {}
    for inc in incomes:
        date = inc[3]  # Assuming 'date_kiritilgan' is at index 3
        income_by_date.setdefault(date, []).append(inc)

    expense_by_date = {}
    for exp in expenses:
        date = exp[3]  # Assuming 'date_kiritilgan' is at index 3
        expense_by_date.setdefault(date, []).append(exp)

    # Get all dates within the range
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    delta = end_dt - start_dt
    dates = [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(delta.days + 1)]

    # Determine if the period exceeds 5 days
    period_length = (end_dt - start_dt).days + 1
    chunk_size = 5
    if period_length > chunk_size:
        # Split dates into 5-day chunks
        date_chunks = [dates[i:i + chunk_size] for i in range(0, len(dates), chunk_size)]
    else:
        date_chunks = [dates]

    # Remove the call to show_pending_approvals
    # await show_pending_approvals(update, context, filial_id, start_date, end_date)

    # Save the date_chunks and current index in user_data for navigation
    context.user_data["report_date_chunks"] = date_chunks
    context.user_data["current_chunk_index"] = 0
    context.user_data["report_filial_id"] = filial_id

    # Display the first chunk
    await display_report_chunk(update, context, filial_id, date_chunks, 0)
    return State.ADMIN_NAVIGATE_DATA



# admin_handlers.py

from telegram import InlineKeyboardButton, InlineKeyboardMarkup

# admin_handlers.py

from telegram import InlineKeyboardButton, InlineKeyboardMarkup

# admin_handlers.py

from telegram import InlineKeyboardButton, InlineKeyboardMarkup

# admin_handlers.py

from telegram import InlineKeyboardButton, InlineKeyboardMarkup

async def display_report_chunk(update: Update, context: ContextTypes.DEFAULT_TYPE, filial_id, date_chunks, index):
    if update.callback_query:
        # If triggered by a callback (navigation), delete the previous report message
        previous_message_id = context.user_data.get("previous_report_message_id")
        if previous_message_id:
            try:
                await update.callback_query.message.delete()
            except:
                pass  # Message might have been already deleted

    chunk_dates = date_chunks[index]

    # Derive current_start_date and current_end_date from the current chunk
    current_start_date = chunk_dates[0]
    current_end_date = chunk_dates[-1]

    msg = f"Filial: {FILIALS.get(filial_id, 'Nomalum')}"

    overall_total_in = 0
    overall_total_out = 0

    for date in chunk_dates:
        incomes = get_incomes_by_filial_and_period(filial_id, date, date)
        expenses = get_expenses_by_filial_and_period(filial_id, date, date)

        if not incomes and not expenses:
            continue  # Skip dates with no data

        msg += f"\n📅 **{date}**\n\n"

        # Calculate daily totals
        daily_total_in = sum(inc[4] + inc[5] + inc[6] + inc[7] + inc[8] for inc in incomes)
        daily_total_out = sum(exp[4] for exp in expenses)
        daily_qolgan = daily_total_in - daily_total_out

        overall_total_in += daily_total_in
        overall_total_out += daily_total_out


        # List incomes with approval status
        if incomes:
            msg += "**Daromadlar:**\n"
            for inc in incomes:
                status = "✅ Tasdiqlangan" if inc[10] == 1 else "❌ Tasdiqlanmagan"
                user = get_user_by_id(inc[1])
                user_name = user[4] if user else "Noma'lum user"
                msg += f"- ID: {inc[0]}, User: {user_name}, Naqt: {inc[4]}, Humo: {inc[5]}, Uzcard: {inc[6]}, Clic: {inc[7]}, Payme: {inc[8]} [{status}]\n\n\n"

        # List expenses with approval status
        if expenses:
            msg += "**Xarajatlar:**\n"
            for exp in expenses:
                status = "✅ Tasdiqlangan" if exp[7] == 1 else "❌ Tasdiqlanmagan"
                user = get_user_by_id(exp[1])
                user_name = user[4] if user else "Noma'lum user"
                msg += f"- ID: {exp[0]}, User: {user_name}, Summa: {exp[4]}, Info: {exp[5]} [{status}]\n"

    # Overall summary for the chunk
    overall_qolgan = overall_total_in - overall_total_out
    msg += f"\n**Umumiy Daromad:** {overall_total_in} so'm\n"
    msg += f"**Umumiy Xarajat:** {overall_total_out} so'm\n"
    msg += f"**Umumiy Qolgan:** {overall_qolgan} so'm\n"

    # Navigation buttons
    buttons = []
    if len(date_chunks) > 1:
        if index > 0:
            buttons.append(InlineKeyboardButton("⬅️ Oldinga", callback_data=f"navigate_back_{index}"))
        if index < len(date_chunks) - 1:
            buttons.append(InlineKeyboardButton("Oldinga ➡️", callback_data=f"navigate_forward_{index}"))

    if buttons:
        keyboard = InlineKeyboardMarkup([buttons])
    else:
        keyboard = None

    sent_message = await update.effective_message.reply_text(
        msg,
        parse_mode='Markdown',
        reply_markup=keyboard
    )

    # Track the sent message ID for future deletion
    context.user_data["previous_report_message_id"] = sent_message.message_id




# admin_handlers.py

# admin_handlers.py
# admin_handlers.py

# admin_handlers.py

async def navigate_forward_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    match = re.match(r"navigate_forward_(\d+)", data)
    if match:
        current_index = int(match.group(1))
        date_chunks = context.user_data.get("report_date_chunks")
        filial_id = context.user_data.get("report_filial_id")
        if date_chunks and filial_id and current_index + 1 < len(date_chunks):
            await display_report_chunk(update, context, filial_id, date_chunks, current_index + 1)
    await query.answer()
    return State.ADMIN_NAVIGATE_DATA

async def navigate_back_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    match = re.match(r"navigate_back_(\d+)", data)
    if match:
        current_index = int(match.group(1))
        date_chunks = context.user_data.get("report_date_chunks")
        filial_id = context.user_data.get("report_filial_id")
        if date_chunks and filial_id and current_index - 1 >= 0:
            await display_report_chunk(update, context, filial_id, date_chunks, current_index - 1)
    await query.answer()
    return State.ADMIN_NAVIGATE_DATA





# admin_handlers.py

# ... existing imports and functions ...

import logging

async def show_today_data(update: Update, context: ContextTypes.DEFAULT_TYPE, filial_id=None):
    today_str = datetime.now().strftime("%Y-%m-%d")
    msg = "=== Bugungi daromad/xarajatlar ===\n"

    if filial_id:
        # Show data for the selected filial only
        incomes = get_incomes_by_filial_and_period(filial_id, today_str, today_str)
        expenses = get_expenses_by_filial_and_period(filial_id, today_str, today_str)
        fname = FILIALS.get(filial_id, "Nomalum Filial")
        msg += f"\nFilial: {fname}\n"

        total_in = sum(row[4] + row[5] + row[6] + row[7] + row[8] for row in incomes)
        total_out = sum(row[4] for row in expenses)  # Assuming 'summa' is at index 4
        qolgan = total_in - total_out
        msg += f"Daromad: {total_in}, Xarajat: {total_out}, Qolgan: {qolgan}\n"

        # Include approval statuses
        msg += "Daromadlar:\n\n"
        for inc in incomes:
            status = "✅ Tasdiqlangan" if inc[10] == 1 else "❌ Tasdiqlanmagan"
            user = get_user_by_id(inc[1])
            user_name = user[4] if user else "Noma'lum user"
            msg += f"ID: {inc[0]}, User: {user_name}, Naqt: {inc[4]}, Humo: {inc[5]}, Uzcard: {inc[6]}, Clic: {inc[7]}, Payme: {inc[8]} [{status}]\n"

        msg += "Xarajatlar:\n\n"
        for exp in expenses:
            status = "✅ Tasdiqlangan" if exp[7] == 1 else "❌ Tasdiqlanmagan"
            user = get_user_by_id(exp[1])
            user_name = user[4] if user else "Noma'lum user"
            msg += f"ID: {exp[0]}, User: {user_name}, Summa: {exp[4]}, Info: {exp[5]} [{status}]\n"
    else:
        # Existing behavior: show data for all filials
        for fid, fname in FILIALS.items():
            incomes = get_incomes_by_filial_and_period(fid, today_str, today_str)
            expenses = get_expenses_by_filial_and_period(fid, today_str, today_str)
            if not incomes and not expenses:
                continue

            total_in = sum(row[4] + row[5] + row[6] + row[7] + row[8] for row in incomes)
            total_out = sum(row[4] for row in expenses)  # Assuming 'summa' is at index 4
            qolgan = total_in - total_out
            msg += f"\nFilial: {fname}\nDaromad: {total_in}, Xarajat: {total_out}, Qolgan: {qolgan}\n\n"

            # Include approval statuses
            msg += "Daromadlar:\n\n"
            for inc in incomes:
                status = "✅ Tasdiqlangan" if inc[10] == 1 else "❌ Tasdiqlanmagan"
                user = get_user_by_id(inc[1])
                user_name = user[4] if user else "Noma'lum user"
                msg += f"ID: {inc[0]}, User: {user_name}, Naqt: {inc[4]}, Humo: {inc[5]}, Uzcard: {inc[6]}, Clic: {inc[7]}, Payme: {inc[8]} [{status}]\n"

            msg += "Xarajatlar:\n\n"
            for exp in expenses:
                status = "✅ Tasdiqlangan" if exp[7] == 1 else "❌ Tasdiqlanmagan"
                user = get_user_by_id(exp[1])
                user_name = user[4] if user else "Noma'lum user"
                msg += f"ID: {exp[0]}, User: {user_name}, Summa: {exp[4]}, Info: {exp[5]} [{status}]\n"

    await update.message.reply_text(msg, reply_markup=admin_panel_keyboard())

    # Fetch pending approvals
    conn = get_connection()
    cursor = conn.cursor()

    # Incomes pending approval
    if filial_id:
        cursor.execute("""
            SELECT * FROM incomes
            WHERE filial_id=? AND date_kiritilgan=? AND confirm_status=1 AND admin_approval=0
        """, (filial_id, today_str,))
    else:
        cursor.execute("""
            SELECT * FROM incomes
            WHERE date_kiritilgan=? AND confirm_status=1 AND admin_approval=0
        """, (today_str,))
    pending_incomes = cursor.fetchall()

    # Expenses pending approval
    if filial_id:
        cursor.execute("""
            SELECT * FROM expenses
            WHERE filial_id=? AND date_kiritilgan=? AND confirm_status=1 AND admin_approval=0
        """, (filial_id, today_str,))
    else:
        cursor.execute("""
            SELECT * FROM expenses
            WHERE date_kiritilgan=? AND confirm_status=1 AND admin_approval=0
        """, (today_str,))
    pending_expenses = cursor.fetchall()
    conn.close()

    # **Qo'shilgan logging qismlari**
    logging.info(f"Pending incomes: {len(pending_incomes)}")
    logging.info(f"Pending expenses: {len(pending_expenses)}")

    if not pending_incomes and not pending_expenses:
        await update.message.reply_text("Hozircha tasdiqlash uchun kutilayotgan ma'lumot yo‘q.",
                                        reply_markup=admin_panel_keyboard())
        return

    # Pending Incomes
    if pending_incomes:
        await update.message.reply_text("Tasdiqlanmagan daromad(lar):")
        for inc in pending_incomes:
            # inc -> (id, user_id, filial_id, date_kiritilgan, naqt, humo, uzkard, clic, payme, confirm_status, admin_approval)
            inc_id = inc[0]
            user_id = inc[1]
            user_row = get_user_by_id(user_id)
            user_name = user_row[4] if user_row else "Noma'lum user"

            s_naqt, s_humo, s_uzkard, s_clic, s_payme = inc[4], inc[5], inc[6], inc[7], inc[8]
            total_sum = s_naqt + s_humo + s_uzkard + s_clic + s_payme

            txt = (f"#DaromadID {inc_id}  (User: {user_name})\n"
                   f"Naqt: {s_naqt}, Humo: {s_humo}, Uzcard: {s_uzkard}, Clic: {s_clic}, Payme: {s_payme}\n"
                   f"Umumiy: {total_sum}")
            await update.message.reply_text(
                txt,
                reply_markup=approval_inline_keyboard("income", inc_id)
            )
    else:
        await update.message.reply_text("Daromad tasdiqlanmagan yo‘q.")

    # Pending Expenses
    if pending_expenses:
        await update.message.reply_text("Tasdiqlanmagan xarajat(lar):")
        for xp in pending_expenses:
            # xp -> (id, user_id, filial_id, date_kiritilgan, summa, info, confirm_status, admin_approval)
            xp_id = xp[0]
            user_id = xp[1]
            user_row = get_user_by_id(user_id)
            user_name = user_row[4] if user_row else "Noma'lum user"

            s_summa = xp[4]
            s_info = xp[5]
            txt = (f"#XarajatID {xp_id} (User: {user_name})\n"
                   f"Summa: {s_summa} so'm, Info: {s_info}")
            await update.message.reply_text(
                txt,
                reply_markup=approval_inline_keyboard("expense", xp_id)
            )
    else:
        await update.message.reply_text("Xarajat tasdiqlanmagan yo‘q.")

    await update.message.reply_text("Quyidagilardan tasdiqlang yoki rad eting.",
                                    reply_markup=admin_panel_keyboard())

    return State.ADMIN_MENU


# admin_handlers.py

# ... mavjud kod ...

async def admin_callback_approval(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Callback: income_approve_XX, income_reject_XX,
              expense_approve_YY, expense_reject_YY
    """
    query = update.callback_query
    data = query.data  # e.g., "income_approve_12"
    await query.answer()

    pattern = r"(income|expense)_(approve|reject)_(\d+)"
    match = re.match(pattern, data)
    if not match:
        await query.message.reply_text("Noma'lum callback.")
        return

    rec_type = match.group(1)   # income yoki expense
    action = match.group(2)     # approve yoki reject
    rec_id = int(match.group(3))

    if rec_type == "income":
        income = get_income_by_id(rec_id)
        if not income:
            await query.message.reply_text("Daromad topilmadi.")
            return

        if action == "approve":
            if income[10] != 0:
                await query.message.reply_text("Bu daromad allaqachon tasdiqlangan yoki rad etilgan.")
                return
            admin_approve_income(rec_id)
            status = "QABUL qilindi"
            # Notify user
            user_row = get_user_by_id(income[1])  # To'g'ri funksiyani chaqirish
            if user_row:
                telegram_id = user_row[1]  # telegram_id joylashuvi: 1-indeks
                await context.bot.send_message(chat_id=telegram_id,
                    text=f"Sizning daromadingiz (ID:{rec_id}) admin tomonidan {status}.")
            await query.message.reply_text(f"Daromad ID:{rec_id} {status}.")

        elif action == "reject":
            if income[10] != 0:
                await query.message.reply_text("Bu daromad allaqachon tasdiqlangan yoki rad etilgan.")
                return
            admin_reject_income(rec_id)
            # Delete the rejected income from the database
            delete_income(rec_id)
            status = "RAD etildi va ma'lumot o'chirildi"
            # Notify user
            user_row = get_user_by_id(income[1])  # To'g'ri funksiyani chaqirish
            if user_row:
                telegram_id = user_row[1]  # telegram_id joylashuvi: 1-indeks
                await context.bot.send_message(chat_id=telegram_id,
                    text=f"Sizning daromadingiz (ID:{rec_id}) admin tomonidan {status}. Iltimos, daromadingizni qayta yuboring.")
            await query.message.reply_text(f"Daromad ID:{rec_id} {status}.")

    elif rec_type == "expense":
        expense = get_expense_by_id(rec_id)
        if not expense:
            await query.message.reply_text("Xarajat topilmadi.")
            return

        if action == "approve":
            if expense[7] != 0:
                await query.message.reply_text("Bu xarajat allaqachon tasdiqlangan yoki rad etilgan.")
                return
            admin_approve_expense(rec_id)
            status = "QABUL qilindi"
            # Notify user
            user_row = get_user_by_id(expense[1])  # To'g'ri funksiyani chaqirish
            if user_row:
                telegram_id = user_row[1]  # telegram_id joylashuvi: 1-indeks
                await context.bot.send_message(chat_id=telegram_id,
                    text=f"Sizning xarajatingiz (ID:{rec_id}) admin tomonidan {status}.")
            await query.message.reply_text(f"Xarajat ID:{rec_id} {status}.")

        elif action == "reject":
            if expense[7] != 0:
                await query.message.reply_text("Bu xarajat allaqachon tasdiqlangan yoki rad etilgan.")
                return
            admin_reject_expense(rec_id)
            # Delete the rejected expense from the database
            delete_expense(rec_id)
            status = "RAD etildi va ma'lumot o'chirildi"
            # Notify user
            user_row = get_user_by_id(expense[1])  # To'g'ri funksiyani chaqirish
            if user_row:
                telegram_id = user_row[1]  # telegram_id joylashuvi: 1-indeks
                await context.bot.send_message(chat_id=telegram_id,
                    text=f"Sizning xarajatingiz (ID:{rec_id}) admin tomonidan {status}. Iltimos, xarajatingizni qayta yuboring.")
            await query.message.reply_text(f"Xarajat ID:{rec_id} {status}.")

    await query.message.reply_text("Quyidagilardan tasdiqlang yoki rad eting.",
                                    reply_markup=admin_panel_keyboard())





# admin_handlers.py

# admin_handlers.py

async def admin_today_view_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    filial_id = context.user_data.get("selected_today_filial")
    if not filial_id:
        await update.message.reply_text("Filial tanlanmagan. Qayta urinib ko'ring.",
                                        reply_markup=admin_panel_keyboard())
        return State.ADMIN_MENU

    today_str = datetime.now().strftime("%Y-%m-%d")
    await show_pending_approvals(update, context, filial_id, today_str, today_str)  # To'g'ri funksiyani chaqirish
    return State.ADMIN_MENU




# Handler for selecting filial during export
async def admin_export_filial_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    match = re.match(r"filial_(\d+)", data)
    if match:
        filial_id = int(match.group(1))
        context.user_data["export_filial_id"] = filial_id
        await query.message.reply_text("Davrni tanlang:", reply_markup=period_select_keyboard())
        await query.answer()
        return State.ADMIN_EXPORT_PERIOD
    else:
        await query.message.reply_text("Filial topilmadi.", reply_markup=admin_panel_keyboard())
        await query.answer()
        return State.ADMIN_MENU

# Handler for selecting period during export
async def admin_export_period_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "Orqaga":
        await update.message.reply_text("Admin menyu", reply_markup=admin_panel_keyboard())
        return State.ADMIN_MENU

    today = datetime.now()
    if text == "Bugungi":
        start_date = today.strftime("%Y-%m-%d")
        end_date = start_date
    elif text == "Kechagi":
        d = today - timedelta(days=1)
        start_date = d.strftime("%Y-%m-%d")
        end_date = start_date
    elif text == "1-Haftalik":
        d = today - timedelta(days=7)
        start_date = d.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif text == "1-Oylik":
        d = today - timedelta(days=30)
        start_date = d.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif text == "Davrni tanlang (manual)":
        await update.message.reply_text("Iltimos, sanani 'YYYY-MM-DD - YYYY-MM-DD' formatida kiriting:",
                                        reply_markup=back_keyboard())
        return State.ADMIN_ENTER_PERIOD_MANUAL
    else:
        await update.message.reply_text("Noma'lum tanlov.",
                                        reply_markup=period_select_keyboard())
        return State.ADMIN_EXPORT_PERIOD

    # Save the selected period
    context.user_data["export_start_date"] = start_date
    context.user_data["export_end_date"] = end_date

    await update.message.reply_text("Qaysi formatda eksport qilmoqchisiz?",
                                    reply_markup=export_inline_keyboard())
    return State.ADMIN_EXPORT_FORMAT

# Handler for manual period entry during export
async def admin_export_enter_period_manual_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "Orqaga":
        await update.message.reply_text("Davrni tanlang:", reply_markup=period_select_keyboard())
        return State.ADMIN_EXPORT_PERIOD

    match = re.match(r"(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})", text)
    if not match:
        await update.message.reply_text("Noto'g'ri format. Masalan: 2025-01-01 - 2025-01-07",
                                        reply_markup=back_keyboard())
        return State.ADMIN_ENTER_PERIOD_MANUAL

    start_date, end_date = match.group(1), match.group(2)
    context.user_data["export_start_date"] = start_date
    context.user_data["export_end_date"] = end_date

    await update.message.reply_text("Qaysi formatda eksport qilmoqchisiz?",
                                    reply_markup=export_inline_keyboard())
    return State.ADMIN_EXPORT_FORMAT

# Callback handler for export format selection
async def export_format_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data  # 'export_pdf' or 'export_excel'
    await query.answer()

    export_format = None
    if data == "export_pdf":
        export_format = "pdf"
    elif data == "export_excel":
        export_format = "excel"
    else:
        await query.message.reply_text("Noma'lum format.")
        return State.ADMIN_EXPORT_FORMAT

    # Retrieve export parameters
    filial_id = context.user_data.get("export_filial_id")
    start_date = context.user_data.get("export_start_date")
    end_date = context.user_data.get("export_end_date")

    if not filial_id or not start_date or not end_date:
        await query.message.reply_text("Export parametrlari to'liq emas. Qayta urinib ko'ring.",
                                        reply_markup=admin_panel_keyboard())
        return State.ADMIN_MENU

    # Fetch approved incomes and expenses
    incomes = get_incomes_by_filial_and_period(filial_id, start_date, end_date)
    expenses = get_expenses_by_filial_and_period(filial_id, start_date, end_date)

    approved_incomes = [inc for inc in incomes if inc[10] == 1]  # admin_approval=1
    approved_expenses = [exp for exp in expenses if exp[7] == 1]  # admin_approval=1

    if export_format == "pdf":
        report_file = generate_pdf_report(filial_id, start_date, end_date, approved_incomes, approved_expenses)
    else:
        report_file = generate_excel_report(filial_id, start_date, end_date, approved_incomes, approved_expenses)

    # Send the file
    try:
        with open(report_file, 'rb') as file:
            await context.bot.send_document(chat_id=query.message.chat_id,
                                           document=file,
                                           filename=os.path.basename(report_file))
    except Exception as e:
        await query.message.reply_text(f"Xatolik yuz berdi: {e}", reply_markup=admin_panel_keyboard())
    finally:
        # Remove the temporary file
        if os.path.exists(report_file):
            os.remove(report_file)

    # Clean up user data
    context.user_data.pop("export_filial_id", None)
    context.user_data.pop("export_start_date", None)
    context.user_data.pop("export_end_date", None)

    # Remove inline buttons
    await query.message.edit_reply_markup(None)
    await query.message.reply_text("Eksport muvaffaqiyatli amalga oshirildi.",
                                    reply_markup=admin_panel_keyboard())

    return State.ADMIN_MENU

# Helper functions to generate reports
def generate_pdf_report(filial_name, start_date, end_date, incomes, expenses):
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm

    report_filename = f"export_report_{filial_name}_{start_date}_to_{end_date}.pdf"
    doc = SimpleDocTemplate(
        report_filename,
        pagesize=landscape(A3),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    elements = []
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Center", alignment=1))

    # Title
    title = Paragraph(
        f"<b>{filial_name} - Hisoboti</b><br/>{start_date} dan {end_date} gacha",
        styles["Title"],
    )
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Prepare data organized by date
    report_data = [
        [
            "Sana",
            "Naqt",
            "Humo",
            "Uzcard",
            "Click",
            "Payme",
            "Daromad",
            "Xarajat",
            "Qolgan Summa",
            "Status",
        ]
    ]

    dates = set()
    for inc in incomes:
        dates.add(inc[3])  # 'date_kiritilgan' assumed to be at index 3
    for exp in expenses:
        dates.add(exp[3])
    sorted_dates = sorted(list(dates))

    for date in sorted_dates:
        naqt = sum(inc[4] for inc in incomes if inc[3] == date)
        humo = sum(inc[5] for inc in incomes if inc[3] == date)
        uzcard = sum(inc[6] for inc in incomes if inc[3] == date)
        click = sum(inc[7] for inc in incomes if inc[3] == date)
        payme = sum(inc[8] for inc in incomes if inc[3] == date)
        daromad = naqt + humo + uzcard + click + payme

        daily_expenses = [exp for exp in expenses if exp[3] == date]
        xarajat_sum = sum(exp[4] for exp in daily_expenses)
        xarajat = f"{xarajat_sum:,} so'm"
        qolgan_summa = daromad - xarajat_sum

        pending_incomes = [inc for inc in incomes if inc[3] == date and inc[10] == 0]
        pending_expenses = [exp for exp in expenses if exp[3] == date and exp[7] == 0]
        status = "✅ Tasdiqlangan" if not pending_incomes and not pending_expenses else "❌ Tasdiqlanmagan"

        report_data.append(
            [
                date,
                f"{naqt:,} so'm",
                f"{humo:,} so'm",
                f"{uzcard:,} so'm",
                f"{click:,} so'm",
                f"{payme:,} so'm",
                f"{daromad:,} so'm",
                xarajat,
                f"{qolgan_summa:,} so'm",
                status,
            ]
        )

    # Create table
    table = Table(
        report_data,
        repeatRows=1,
        hAlign="CENTER",
        colWidths=[
            30 * mm,
            40 * mm,
            40 * mm,
            40 * mm,
            40 * mm,
            40 * mm,
            40 * mm,
            50 * mm,
            40 * mm,
            30 * mm,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
            ]
        )
    )

    for i in range(1, len(report_data)):
        bg_color = colors.lightgrey if i % 2 == 0 else colors.whitesmoke
        table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), bg_color)]))

    elements.append(table)

    doc.build(elements)
    return report_filename

def generate_excel_report(filial_id, start_date, end_date, incomes, expenses):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    report_filename = f"export_report_{filial_id}_{start_date}_to_{end_date}.xlsx"

    # Prepare data organized by date
    report_data = []

    # Get unique dates
    dates = set()
    for inc in incomes:
        dates.add(inc[3])  # 'date_kiritilgan' assumed to be at index 3
    for exp in expenses:
        dates.add(exp[3])
    sorted_dates = sorted(list(dates))

    for date in sorted_dates:
        # Sum incomes by method
        naqt = sum(inc[4] for inc in incomes if inc[3] == date)
        humo = sum(inc[5] for inc in incomes if inc[3] == date)
        uzcard = sum(inc[6] for inc in incomes if inc[3] == date)
        click = sum(inc[7] for inc in incomes if inc[3] == date)
        payme = sum(inc[8] for inc in incomes if inc[3] == date)
        daromad = naqt + humo + uzcard + click + payme

        # Sum expenses
        xarajat = sum(exp[4] for exp in expenses if exp[3] == date)

        # Calculate Qolgan summa
        qolgan_summa = daromad - xarajat

        # Determine status
        pending_incomes = [inc for inc in incomes if inc[3] == date and inc[10] == 0]
        pending_expenses = [exp for exp in expenses if exp[3] == date and exp[7] == 0]
        if not pending_incomes and not pending_expenses:
            status = "Tasdiqlangan"
        else:
            status = "Tasdiqlanmagan"

        # Append row
        report_data.append([
            date,
            f"{naqt:,} so'm",
            f"{humo:,} so'm",
            f"{uzcard:,} so'm",
            f"{click:,} so'm",
            f"{payme:,} so'm",
            f"{daromad:,} so'm",
            f"{xarajat:,} so'm",
            f"{qolgan_summa:,} so'm",
            status
        ])

    # Create DataFrame
    df_report = pd.DataFrame(report_data, columns=[
        "Sana", "Naqt", "Humo", "Uzcard", "Click", "Payme", "Daromad", "Xarajat", "Qolgan Summa", "Status"
    ])

    # Create Excel file with OpenPyXL
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    # Write headers
    headers = list(df_report.columns)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_num, row_data in enumerate(report_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply conditional formatting to the Status column
            if col_num == len(row_data):  # 'Status' column
                if cell_value == "Tasdiqlangan":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red

    # Auto-adjust column widths
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = max(len(str(cell.value) or "") for cell in column_cells)
        adjusted_width = max_length + 2
        sheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Save the workbook
    workbook.save(report_filename)
    return report_filename
